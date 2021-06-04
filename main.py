'''
       @author: joydeep ghosh 99004391
       @description :main module implementing all the classes
       @date : 3-06-2021
'''

from openpyxl import load_workbook, Workbook

class WorkbookIo():
    '''
       @author: joydeep ghosh 99004391
       @description :WorkbookIo class defining classes
       @date : 3-06-2021
    '''
    def readworkbook(self):
        '''
        @author: joydeep ghosh 99004391
        @description :method implementing read workbook
        @date : 3-06-2021
        '''
        try:
            workbookload = load_workbook(filename='input_data.xlsx')
            sheet_names = workbookload.sheetnames
            data = {}
            for name in sheet_names:
                databuff = []
                worksheet = workbookload[name]
                cellrange = worksheet['A1':'U17']
                row_list = []
                for row in cellrange:
                    row1 = []
                    for col in row:
                        row1.append(str(col.value))
                    row_list.append(row1)
                databuff.append(row_list)
                data[name] = databuff
            return data
        except IOError as ex:
            print("file reading error", ex)
            return None
    def write_data(self, data):
        '''
        @author: joydeep ghosh 99004391
        @description :method implementing writing data to workbook
        @date : 3-06-2021
        '''
        try:
            wb1 = Workbook()
            destination_file_name = 'output_data.xlsx'
            ws1 = wb1.active
            ws1.title = 'selected'
            ws1['A1'] = 'data'
            list2 = [['psno']]
            for i in range(1, 2):
                for j in range(1, 2):
                    ws1.cell(column=j, row=i, value=list2[i - 1][j - 1])
            for psn in data.keys():
                for row in range(2, 3):
                    ws1.cell(row=row, column=1, value=psn)
                    for col in range(2, 22):
                        ws1.cell(row=row, column=col, value=data[psn][0][col - 2])
                    # print(data[psn][0][1])
            wb1.save(filename=destination_file_name)
        except IOError as ex:
            print("file writing error", ex)
        return 0



class WorkBook(WorkbookIo):
    '''
    @author: joydeep ghosh 99004391
    @description : Workbook class inheriting from WorkbookIo
    @date : 3-06-2021

    '''
    def getdata(self, data, psno):
        '''
        @author: joydeep ghosh 99004391
        @description :method implementing getting all data
        @date : 3-06-2021
        '''
        try:
            total_data = {}
            total_data_buffer = []

            for i in data.keys():
                data_buff = data[i]

                for k in range(1, len(data_buff[0])):

                    psn = data_buff[0][k][0]
                    # print("yyy",data_buff[0][k])
                    if psn == psno:
                        total_data_buffer.append(data_buff[0][k][1:])
                total_data[psno] = total_data_buffer
        except ValueError as ex:
            print("error in reading data", ex)
        return total_data
    def getdata_select(self, data, psno, dataselect):
        '''
         @author: joydeep ghosh 99004391
         @description :method implementing getting selected data
         @date : 3-06-2021
        '''
        try:
            total_data = {}
            total_data_buffer = []
            try:
                if float(psno) < 99004390.0 or float(psno)>99004405.0:
                    raise ValueError("Enter valid psno")
            except ValueError as ex:
                print(ex)
            try:
                if dataselect not in ['games','tvseries','movies','books','football']:
                    raise ValueError("Enter valid choice of data")
            except ValueError as ex:
                print(ex)
            for i in data.keys():
                if i == dataselect:
                    data_buff = data[i]

                    for k in range(1, len(data_buff[0])):

                        psn = data_buff[0][k][0]
                        # print("yyy",data_buff[0][k])
                        if psn == psno:
                            total_data_buffer.append(data_buff[0][k][1:])
                    total_data[psno] = total_data_buffer
        except ValueError as ex:
            print("error in reading selected data", ex)
        return total_data

if __name__ == '__main__':
    PS_NO = str(float(input("enter psno from \n99004390 \n99004391 \n99004392 \n99004393 \n99004394 \n99004395 \n99004396 \n99004397 \n99004398 \n99004399 \n99004400 \n99004401 \n99004402 \n99004403 \n99004404 \n99004405\nyour choice psno:")))
    data_to_be_selected = input("enter data to be selected  from \ngames \ntvseries \nmovies \nbooks \nfootball\n your choice of data:")
    workbook = WorkBook()
    Data_Read = workbook.readworkbook()
    data_by_select = workbook.getdata_select(Data_Read, PS_NO, data_to_be_selected)
    workbook.write_data(data_by_select)
